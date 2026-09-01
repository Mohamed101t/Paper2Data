import csv
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.services.export_mapping import ExportMapping
from domain.entities.field import Field
from domain.entities.field_type import FieldType
from domain.entities.record import Record


class ExportService:
    """Export records using field definitions rather than spreadsheet guessing."""

    @staticmethod
    def export_to_excel(fields: List[Field], records: List[Record], file_path: str) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet.freeze_panes = "A2"

        headers = ["ID", "Created At"] + [field.name for field in fields]
        worksheet.append(headers)

        header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for record in records:
            value_map = {value.field_id: value.value for value in record.values}
            row_number = worksheet.max_row + 1

            id_cell = worksheet.cell(row=row_number, column=1, value=record.id)
            id_cell.number_format = ExportMapping.INTEGER_FORMAT

            created_value = ExportService._parse_created_at(record.created_at)
            created_cell = worksheet.cell(row=row_number, column=2, value=created_value)
            if isinstance(created_value, datetime):
                created_cell.number_format = ExportMapping.DATETIME_FORMAT

            for column_index, field in enumerate(fields, start=3):
                field_type = FieldType.normalize(field.field_type)
                raw_value = record.id if field_type == FieldType.AUTO_NUMBER else value_map.get(field.id, "")
                mapped = ExportMapping.to_excel(field_type, raw_value)
                cell = worksheet.cell(row=row_number, column=column_index, value=mapped.value)
                if mapped.number_format:
                    cell.number_format = mapped.number_format
                if mapped.force_text:
                    # Explicit string type prevents user text beginning with "="
                    # from becoming an XLSX formula.
                    cell.data_type = "s"

        worksheet.auto_filter.ref = worksheet.dimensions
        ExportService._fit_columns(worksheet)
        workbook.save(file_path)

    @staticmethod
    def export_to_csv(fields: List[Field], records: List[Record], file_path: str) -> None:
        headers = ["ID", "Created At"] + [field.name for field in fields]
        with open(file_path, "w", encoding="utf-8-sig", newline="") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(headers)
            for record in records:
                value_map = {value.field_id: value.value for value in record.values}
                row = [record.id, ExportService._csv_created_at(record.created_at)]
                for field in fields:
                    field_type = FieldType.normalize(field.field_type)
                    raw_value = record.id if field_type == FieldType.AUTO_NUMBER else value_map.get(field.id, "")
                    row.append(ExportMapping.to_csv(field_type, raw_value))
                writer.writerow(row)

    @staticmethod
    def _parse_created_at(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value
        text = str(value).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        return text

    @staticmethod
    def _csv_created_at(value) -> str:
        parsed = ExportService._parse_created_at(value)
        if isinstance(parsed, datetime):
            return parsed.strftime("%Y-%m-%d %H:%M:%S")
        return "" if parsed is None else str(parsed)

    @staticmethod
    def _fit_columns(worksheet) -> None:
        for column_index, column_cells in enumerate(worksheet.columns, start=1):
            max_length = max(
                (len(str(cell.value)) for cell in column_cells if cell.value is not None),
                default=0,
            )
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 10), 40)
