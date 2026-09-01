import csv

import pytest
from openpyxl import load_workbook

from core.services.export_service import ExportService
from data.repositories.field_repository_impl import FieldRepositoryImpl
from data.repositories.project_repository_impl import ProjectRepositoryImpl
from data.repositories.record_repository_impl import RecordRepositoryImpl
from domain.entities.field_type import FieldType
from presentation.viewmodels.field_viewmodel import FieldViewModel
from presentation.viewmodels.project_viewmodel import ProjectViewModel
from presentation.viewmodels.record_viewmodel import RecordViewModel
from presentation.viewmodels.records_list_viewmodel import RecordsListViewModel


pytestmark = pytest.mark.e2e


def test_complete_project_to_excel_and_csv_journey(quality_db_service, tmp_path):
    project_repo = ProjectRepositoryImpl(quality_db_service)
    field_repo = FieldRepositoryImpl(quality_db_service)
    record_repo = RecordRepositoryImpl(quality_db_service)

    project_vm = ProjectViewModel(project_repo)
    field_vm = FieldViewModel(field_repo)
    record_vm = RecordViewModel(record_repo)
    records_vm = RecordsListViewModel(record_repo, field_repo)

    project_vm.create_project("Customer survey", "E2E")
    project = project_repo.get_all_projects()[0]

    field_vm.set_project(project.id)
    field_vm.add_field("Name", FieldType.SHORT_TEXT, True)
    field_vm.add_field("Age", FieldType.INTEGER, True)
    field_vm.add_field("Phone", FieldType.PHONE_NUMBER, False)
    field_vm.add_field("Success", FieldType.PERCENTAGE, False)
    field_vm.add_field("Active", FieldType.YES_NO, False)
    fields = field_repo.get_fields_by_project(project.id)
    assert len(fields) == 5

    by_name = {field.name: field for field in fields}
    form_data = {
        by_name["Name"].id: "Mohamed",
        by_name["Age"].id: "25",
        by_name["Phone"].id: "00123456789",
        by_name["Success"].id: "85%",
        by_name["Active"].id: "نعم",
    }
    assert record_vm.save_record(project.id, fields, form_data) is True

    received = []
    records_vm.records_loaded.connect(lambda loaded_fields, records: received.append((loaded_fields, records)))
    records_vm.set_project(project.id)
    assert len(received) == 1
    loaded_fields, records = received[0]
    assert len(records) == 1

    xlsx_path = tmp_path / "customer_survey.xlsx"
    csv_path = tmp_path / "customer_survey.csv"
    ExportService.export_to_excel(loaded_fields, records, str(xlsx_path))
    ExportService.export_to_csv(loaded_fields, records, str(csv_path))

    workbook = load_workbook(xlsx_path, data_only=False)
    sheet = workbook["Data"]
    assert sheet.max_row == 2
    assert sheet["C2"].value == "Mohamed"
    assert sheet["D2"].value == 25
    assert sheet["E2"].value == "00123456789"
    assert sheet["E2"].data_type == "s"
    assert sheet["F2"].value == pytest.approx(0.85)
    assert sheet["G2"].value is True

    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    assert rows[1][2:] == ["Mohamed", "25", "00123456789", "85", "yes"]

    records_vm.delete_record(records[0].id)
    assert record_repo.get_records_by_project(project.id) == []
